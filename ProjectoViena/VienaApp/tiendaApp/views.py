from django.forms import ValidationError
from django.shortcuts import render, redirect, get_object_or_404
from tiendaApp.forms import ProductoForm, NewUserForm, CategoriaForm, MesaForm, DetalleComandaForm, UserUpdateForm, UserProfileUpdateForm
from tiendaApp.models import Productos, Categoria, Comanda, DetalleComanda, UserProfile, User
from django.db.models import Q
from django.contrib.auth import login
from django.contrib import messages
from openpyxl import Workbook
import datetime
from django.http import HttpResponse
from django.utils.timezone import make_naive, get_default_timezone, is_naive
from collections import Counter
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from tiendaApp.serializers import ComandaSerializer
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view
from rest_framework_jwt.settings import api_settings
from django.contrib.auth.decorators import user_passes_test, login_required

# Create your views here.

def base(request):
    return render(request, 'base.html')

def es_encargado(user):
    return user.is_superuser or (hasattr(user, 'userprofile') and user.userprofile.cargo == 'Encargado')

def inicio(request):
    return render(request, 'inicio.html')

@login_required
@user_passes_test(es_encargado)
def modificar(request):
    categorias = Categoria.objects.all()  # Obtiene todas las categorías
    return render(request, 'modificar.html', {'categorias': categorias})

@login_required
@user_passes_test(es_encargado)
def ingresoproducto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                form.save()
                return redirect('listaproductos')
            except ValidationError as e:
                # Manejar el error si hay un nombre duplicado
                form.add_error(None, e)
    else:
        form = ProductoForm()
    return render(request, 'ingresoproducto.html', {'form': form})


def listaproductos(request):
    query = request.GET.get('q')
    productos = Productos.objects.all()

    if query:
        # Filtra los productos que contienen la consulta en el nombre o la descripción
        productos = productos.filter(Q(nombre__icontains=query))

    return render(request, 'listaproductos.html', {'productos': productos, 'query': query})

def productos_por_categoria(request, categoria_id):
    productos = Productos.objects.filter(categoria_id=categoria_id)
    categoria = Categoria.objects.get(id=categoria_id)
    return render(request, 'productos_por_categoria.html', {'productos': productos, 'categoria': categoria})

def detalles_categoria(request, categoria_id):
    productos = Productos.objects.filter(categoria_id=categoria_id)
    return render(request, 'detalles_categoria.html', {'productos': productos})

@login_required
@user_passes_test(es_encargado)
def producto_detalle(request, producto_id):
    producto = get_object_or_404(Productos, id=producto_id)

    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('listaproductos')
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'producto_detalle.html', {'form': form, 'producto': producto})
    

@login_required
@user_passes_test(es_encargado)    
def editar_producto(request, producto_id):
    producto = get_object_or_404(Productos, id=producto_id)
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('producto_detalle', producto_id=producto.id)
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'editar_producto.html', {'form': form})

@login_required
@user_passes_test(es_encargado)
def eliminar_producto(request, producto_id):
    # Obtiene el objeto del producto o muestra un error 404 si no existe
    producto = get_object_or_404(Productos, id=producto_id)
    
    if request.method == 'POST':
        # Si la solicitud es POST, elimina el producto
        producto.delete()
        # Redirige a la página de inicio o a donde desees después de la eliminación
        return redirect('inicio')
    
    # Renderiza una página de confirmación de eliminación (puedes crearla si lo deseas)
    return render(request, 'confirmar_eliminacion.html', {'producto': producto})

@login_required
@user_passes_test(es_encargado)
def register_request(request):
    if request.method == "POST":
        form = NewUserForm(request.POST)
        if form.is_valid():
            # Crear el usuario
            user = form.save(commit=False)
            user.save()
            
            # Crear el perfil de usuario asociado
            UserProfile.objects.create(
                user=user,
                nombre=form.cleaned_data['nombre'],
                apellido=form.cleaned_data['apellido'],
                cargo=form.cleaned_data['cargo'],
                rut=form.cleaned_data['rut']
            )

            # Mensajes y redirección
            messages.success(request, "Registro exitoso.")
            return redirect("inicio")
        else:
            # Manejar el caso en que el formulario no es válido
            messages.error(request, "Por favor, corrija los errores")
    else:
        form = NewUserForm()

    return render(request, "registration/registro.html", {"register_form": form})


@login_required
def crear_comanda(request):
    if request.method == 'POST':
        comanda = Comanda.objects.create(usuario_asignado=request.user)

        for key, value in request.POST.items():
            if key.startswith('cantidad_') and value:
                producto_id = key.split('_')[1]
                cantidad = int(value)
                producto = get_object_or_404(Productos, id=producto_id)
                DetalleComanda.objects.create(comanda=comanda, producto=producto, cantidad=cantidad)

        return redirect('confirmar_comanda', comanda_id=comanda.id)

    else:
        # Asumiendo que comestibles, bebestibles y bebidasalcoholicas son tus categorías de productos
        comestibles = Productos.objects.filter(categoria=1)
        bebestibles = Productos.objects.filter(categoria=2)
        bebidasalcoholicas = Productos.objects.filter(categoria=3)

        return render(request, 'crear_comanda.html', {
            'comestibles': comestibles,
            'bebestibles': bebestibles,
            'bebidasalcoholicas': bebidasalcoholicas
        })

def vista_cocina(request):
    comandas = Comanda.objects.prefetch_related('detallecomanda_set').select_related('usuario_asignado__userprofile').all().order_by('-fecha_creacion')
    return render(request, 'vista_cocina.html', {'comandas': comandas})


@login_required
@user_passes_test(es_encargado)
def generar_informe_ventas(request):
    if request.method == 'POST':
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')

        # Convertir las fechas de string a objetos datetime
        fecha_inicio = datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin = datetime.datetime.strptime(fecha_fin, '%Y-%m-%d').date()

        # Crear un libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe de Ventas"

        # Añadir encabezados a la hoja de Excel
        columnas = ['ID Comanda', 'Fecha', 'Total']
        ws.append(columnas)

        # Filtrar comandas por el rango de fechas
        comandas = Comanda.objects.filter(fecha_creacion__range=[fecha_inicio, fecha_fin])

        # Añadir datos de comandas a la hoja
        for comanda in comandas:
            # Convertir la fecha a zona horaria neutral
            fecha_sin_tz = make_naive(comanda.fecha_creacion)
            ws.append([comanda.id, fecha_sin_tz, comanda.total])

        # Configurar la respuesta para descargar el archivo de Excel
        response = HttpResponse(content_type='application/vnd.ms-excel')
        nombre_archivo = f"Informe_Ventas_{fecha_inicio}_{fecha_fin}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        # Guardar el libro de trabajo en la respuesta
        wb.save(response)

        return response

    # Si no es un POST, renderizar el formulario
    return render(request, 'generar_informe.html')

@login_required
@user_passes_test(es_encargado)
def informe_producto(request):
    if request.method == 'POST':
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')

        # Convertir las fechas de string a objetos datetime
        fecha_inicio = datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin = datetime.datetime.strptime(fecha_fin, '%Y-%m-%d')

        # Verifica si las fechas son naive y, si no, conviértelas
        if not is_naive(fecha_inicio):
            fecha_inicio = make_naive(fecha_inicio, get_default_timezone())
        if not is_naive(fecha_fin):
            fecha_fin = make_naive(fecha_fin, get_default_timezone())

        # Obtener todas las comandas en el rango de fechas
        comandas = Comanda.objects.filter(fecha_creacion__range=[fecha_inicio, fecha_fin])

        # Contabilizar los productos vendidos
        contador_productos = Counter()
        for comanda in comandas:
            for detalle in comanda.detallecomanda_set.all():
                contador_productos[detalle.producto] += detalle.cantidad

        # Crear un libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos Vendidos"

        # Crear un libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos Vendidos"

        # Añadir encabezados a la hoja de Excel
        ws.append(['Producto', 'Descripción', 'Precio', 'Cantidad Vendida', 'Total'])

        # Añadir datos de los productos vendidos
        for producto, cantidad in contador_productos.most_common():
            total = producto.precio * cantidad
            ws.append([
                producto.nombre, 
                producto.descripcion, 
                producto.precio, 
                cantidad,
                total  # Total calculado
            ])

        # Configurar la respuesta para descargar el archivo de Excel
        response = HttpResponse(content_type='application/vnd.ms-excel')
        nombre_archivo = f"Productos_Vendidos_{fecha_inicio}_{fecha_fin}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        # Guardar el libro de trabajo en la respuesta
        wb.save(response)

        return response

    # Si no es un POST, renderizar el formulario
    return render(request, 'informe_producto.html')

@login_required
@user_passes_test(es_encargado)
def seleccionar_informes(request):
    return render(request, 'seleccionar_informes.html')

@login_required
def comanda_exitosa(request, comanda_id):
    comanda = get_object_or_404(Comanda, id=comanda_id)
    detalles_comanda = comanda.detallecomanda_set.all()
    total_comanda = sum(detalle.subtotal for detalle in detalles_comanda)

    return render(request, 'comanda_exitosa.html', {
        'comanda': comanda,
        'detalles_comanda': detalles_comanda,
        'total_comanda': total_comanda 
    })


@login_required
def confirmar_comanda(request, comanda_id):
    comanda = get_object_or_404(Comanda, id=comanda_id)
    detalles_comanda = comanda.detallecomanda_set.all()
    total_comanda = sum(detalle.subtotal for detalle in detalles_comanda)

    if request.method == 'POST':
        comanda.confirmada = True
        comanda.save()
        return redirect('comanda_exitosa', comanda_id=comanda.id)

    return render(request, 'confirmar_comanda.html', {
        'comanda': comanda,
        'detalles_comanda': detalles_comanda,
        'total_comanda': total_comanda
    })

@login_required
@user_passes_test(es_encargado)
def lista_trabajadores(request):
    trabajadores = UserProfile.objects.all()
    return render(request, 'lista_trabajadores.html', {'trabajadores': trabajadores})

@login_required
@user_passes_test(es_encargado)
def crear_categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('modificar')
    else:
        form = CategoriaForm()
    return render(request, 'crear_categoria.html', {'form': form})

def lista_categoria(request):
    query = request.GET.get('q')
    categorias = Categoria.objects.all()  # Cambia el nombre de la variable a 'categorias'

    if query:
        categorias = categorias.filter(Q(nombre__icontains=query))

    return render(request, 'lista_categoria.html', {'categorias': categorias, 'query': query})

@login_required
@user_passes_test(es_encargado)
def editar_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)

    if request.method == 'POST':
        form = CategoriaForm(request.POST, request.FILES, instance=categoria)
        if form.is_valid():
            form.save()
            # Redirige a donde consideres adecuado después de editar
            return redirect('lista_categoria')
    else:
        form = CategoriaForm(instance=categoria)

    return render(request, 'editar_categoria.html', {'form': form, 'categoria': categoria})

@login_required
@user_passes_test(es_encargado)
def detalles_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    return render(request, 'detalle_categoria.html', {'categoria': categoria})

@login_required
@user_passes_test(es_encargado)
def eliminar_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)

    if request.method == 'POST':
        categoria.delete()
        messages.success(request, "Categoría eliminada con éxito.")
        return redirect('lista_categoria')  # Redirige a la lista de categorías

    return render(request, 'confirmar_eliminacion_categoria.html', {'categoria': categoria})


@login_required
@user_passes_test(es_encargado)
def modificar_usuario(request, user_id):
    user = get_object_or_404(User, id=user_id)
    
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        # Si no existe un UserProfile, maneja la situación. Por ejemplo:
        messages.error(request, "No existe un perfil para este usuario.")
        return redirect('lista_usuarios')  # Redirige a alguna página relevante


    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=user)
        profile_form = UserProfileUpdateForm(request.POST, instance=user_profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, 'El usuario ha sido actualizado.')
            return redirect('lista_usuarios')  # Redirige a donde sea apropiado
    else:
        user_form = UserUpdateForm(instance=user)
        profile_form = UserProfileUpdateForm(instance=user_profile)

    context = {
        'user_form': user_form,
        'profile_form': profile_form
    }
    return render(request, 'modificar_usuario.html', context)

@login_required
@user_passes_test(es_encargado)
def lista_usuarios(request):
    usuarios = User.objects.all()
    return render(request, 'lista_usuarios.html', {'usuarios': usuarios})

@login_required
@user_passes_test(es_encargado)
def eliminar_usuario(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        user.delete()  # Esto también eliminará el UserProfile debido a la relación en cascada
        messages.success(request, "Usuario eliminado con éxito.")
        return redirect('lista_usuarios')  # Redirige a la lista de usuarios

    return render(request, 'eliminar_usuario.html', {'user': user})

@login_required
@user_passes_test(es_encargado)
def detalles_usuario(request, user_id):
    user = get_object_or_404(User, id=user_id)
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        user_profile = None
        messages.warning(request, "Este usuario no tiene un perfil asociado.")

    context = {
        'user': user,
        'user_profile': user_profile,
    }
    return render(request, 'detalles_usuario.html', context)